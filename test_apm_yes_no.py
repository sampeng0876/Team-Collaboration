import openpyxl
from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook('test.xlsx')

# Select sheet2
sheet = workbook['APM']  # Change the sheet name accordingly

# Define the data
list1 = ['', 'EMCU8377066', '143357136749', '', 'DEPARTED', 'EGL', 'EVER LASTING', '05/10/23 08:00', '05/11/23 14:16', 'COMMUNITY - OUT', 'RELEASED', 'RELEASED', '', '', '', '40/GP/96', '28065.0', '', '05/16/23 08:00', '05/16/23 10:52', '', 'MEDU7539050', 'RX067313', '', 'YARD', 'MSC', 'MAERSK EDIRNE', '05/11/23 23:35', '05/15/23 13:37', 'Yard Grounded (TIS-2E18581)', 'RELEASED', 'RELEASED', '', '05/19/23', '', '40/GP/96', '31733.0', '', '05/19/23 08:00', '']
list2 = ['ready', 'not-ready']

# Replacing values in list2
list2 = ['Yes' if item == 'ready' else 'No' for item in list2]

# Define the starting row and column
start_row = 2
start_col = 1

# Iterate through the data and store it in the worksheet
for i, value in enumerate(list1):
    col = start_col + (i % 20)
    row = start_row + (i // 20)

    # Check if the 4th, 24th, 44th, 64th, and so on elements are empty
    if i % 20 == 3 and value == '':
        value = list2[i // 20 % len(list2)]

    sheet.cell(row=row, column=col).value = value

# Save the workbook
workbook.save('test.xlsx')


# list2 = ['Yes' if item == 'ready' else 'No' for item in list2]

# print(list2)

# index = 3  # Start index to check
# step = 20  # Step size to check subsequent elements

# for i in range(index, len(list1), step):
#     if list1[i] == '':
#         list1[i] = list2[(i // step) % len(list2)]

# print(list1)
