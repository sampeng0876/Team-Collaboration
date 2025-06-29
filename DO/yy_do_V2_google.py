import os
from openpyxl import load_workbook
from datetime import date
from googleapiclient.discovery import build
from google.oauth2 import service_account

def extract_merged_cell_values(file_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path, data_only=True)  # Open with data_only=True to get calculated values
    sheet = wb.active

    # Dictionary to store extracted values in the specified order
    extracted_values = {}

    # Extract merged cell values
    merged_cells = sheet.merged_cells
    for merged_cell in merged_cells:
        if merged_cell.min_row == 4 and merged_cell.max_row == 4:
            if merged_cell.min_col == 7 and merged_cell.max_col == 8:  # G4 and H4 (DATE)
                extracted_values['DATE'] = str(date.today())  # Set to today's date (G4 and H4 merged cells)
                # Handle ETA (G4 and H4 merged cells)
                eta_value = sheet.cell(row=4, column=7).value  # Access the merged cell value
                # Handle the value properly if it's a string or datetime
                if isinstance(eta_value, str):
                    extracted_values['ETA'] = eta_value.strip()  # Strip whitespace if it's a string
                elif isinstance(eta_value, date):
                    extracted_values['ETA'] = eta_value.strftime('%m/%d/%Y')  # Format as string if it's a datetime

            elif merged_cell.min_col == 3 and merged_cell.max_col == 4:  # C4 and D4 (PICK UP)
                extracted_values['PICK UP'] = sheet.cell(row=4, column=3).value

        if merged_cell.min_row == 2 and merged_cell.max_row == 2:
            if merged_cell.min_col == 10 and merged_cell.max_col == 11:  # J2 and K2 (CTNR#)
                ctnr_value = sheet.cell(row=2, column=10).value
                # Strip whitespace if present
                extracted_values['CTNR#'] = ctnr_value.strip() if ctnr_value else None

    # Extract individual cell values in the specified order
    extracted_values['DELIVERY'] = sheet.cell(row=6, column=10).value  # J6 (DELIVERY)
    extracted_values['CUSTOMER'] = "YY"  # Always return "YY"
    extracted_values['TYPE'] = sheet.cell(row=4, column=9).value  # I4 (TYPE)
    extracted_values['WEIGHT'] = sheet.cell(row=7, column=7).value  # G7 (WEIGHT)
    extracted_values['REMARK'] = "N/A"  # Always return "N/A"

    # Ensure the extracted values are in the specified order
    ordered_values = [
        extracted_values.get('DATE'),
        extracted_values.get('CTNR#'),
        extracted_values.get('DELIVERY'),
        extracted_values.get('CUSTOMER'),
        extracted_values.get('TYPE'),
        extracted_values.get('WEIGHT'),
        extracted_values.get('REMARK'),
        extracted_values.get('ETA'),
        extracted_values.get('PICK UP')
    ]

    return ordered_values

def process_excel_to_google_sheet(pdf_paths, google_sheet_id, sheet_name):
    # Google Sheets API setup
    SERVICE_ACCOUNT_FILE = 'service_account.json'
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)

    for file_path in pdf_paths:
        print(f"Processing file: {file_path}")
        extracted_data = extract_merged_cell_values(file_path)
        
        # Print the extracted data before updating Google Sheet
        print("Extracted Data (Before writing to Google Sheet):")
        print(extracted_data)
        
        values = [extracted_data]
        body = {'values': values}

        # Find the next available row
        result = service.spreadsheets().values().get(spreadsheetId=google_sheet_id, range=sheet_name).execute()
        next_row = len(result['values']) + 1

        # Write values to the sheet
        service.spreadsheets().values().update(
            spreadsheetId=google_sheet_id,
            range=sheet_name + '!A' + str(next_row),
            valueInputOption='USER_ENTERED',
            body=body
        ).execute()
        print(f"Data from {file_path} written to Google Sheet.")

def scan_and_process_files():
    excel_paths = []
    directory = r'K:\My Drive\Company\FaCai\DO\YY'  # Directory containing Excel files
    for filename in os.listdir(directory):
        # Skip temporary files and hidden files (those that start with ~$)
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            excel_paths.append(os.path.join(directory, filename))

    if excel_paths:
        google_sheet_id = '12lnRmQoBsITIYTQPEGYdHGVNUkoPPFQEhx5HaC3JTJQ'  # Replace with your Google Sheet ID
        sheet_name = 'RECORD_DO'  # Replace with your sheet name
        process_excel_to_google_sheet(excel_paths, google_sheet_id, sheet_name)  # Process and store data in Google Sheets
    else:
        print("No Excel files found in the directory!")

if __name__ == "__main__":
    scan_and_process_files()
