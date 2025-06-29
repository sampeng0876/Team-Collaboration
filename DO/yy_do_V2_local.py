import os
from openpyxl import load_workbook
from datetime import date

def extract_merged_cell_values(file_path):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path, data_only=True)  # Open with data_only=True to get calculated values
    sheet = wb.active

    # Dictionary to store extracted values
    extracted_values = {}

    # Extract merged cell values
    merged_cells = sheet.merged_cells
    for merged_cell in merged_cells:
        if merged_cell.min_row == 2 and merged_cell.max_row == 2:
            if merged_cell.min_col == 10 and merged_cell.max_col == 11:  # J2 and K2 (CTNR#)
                ctnr_value = sheet.cell(row=2, column=10).value
                # Strip whitespace if present
                extracted_values['CTNR#'] = ctnr_value.strip() if ctnr_value else None

        if merged_cell.min_row == 4 and merged_cell.max_row == 4:
            if merged_cell.min_col == 7 and merged_cell.max_col == 8:  # G4 and H4 (DATE and ETA)
                # Handle DATE (same as before)
                extracted_values['DATE'] = str(date.today())  # Get today's date
                # Handle ETA (G4 and H4 merged cells)
                eta_value = sheet.cell(row=4, column=7).value  # Access the merged cell value
                # Handle the value properly if it's a string or datetime
                if isinstance(eta_value, str):
                    extracted_values['ETA'] = eta_value.strip()  # Strip whitespace if it's a string
                elif isinstance(eta_value, date):
                    extracted_values['ETA'] = eta_value.strftime('%m/%d/%Y')  # Format as string if it's a datetime

            elif merged_cell.min_col == 3 and merged_cell.max_col == 4:  # C4 and D4 (PICK UP)
                extracted_values['PICK UP'] = sheet.cell(row=4, column=3).value

        if merged_cell.min_row == 6 and merged_cell.max_row == 6:
            if merged_cell.min_col == 7 and merged_cell.max_col == 8:  # G7 and H7 (ETA)
                # This is no longer needed, since ETA is now from G4:H4
                pass

    # Extract individual cell values
    extracted_values['DELIVERY'] = sheet.cell(row=6, column=10).value  # J6 (DELIVERY)
    extracted_values['CUSTOMER'] = "YY"  # Always return "YY"
    extracted_values['TYPE'] = sheet.cell(row=4, column=9).value  # I4 (TYPE)
    extracted_values['WEIGHT'] = sheet.cell(row=7, column=7).value  # G7 (WEIGHT)
    extracted_values['REMARK'] = "N/A"  # Always return "N/A"

    # Print the extracted values in the specified order
    print("Extracted Values:")
    print(f"DATE: {extracted_values.get('DATE')}")
    print(f"CTNR#: {extracted_values.get('CTNR#')}")
    print(f"DELIVERY: {extracted_values.get('DELIVERY')}")
    print(f"CUSTOMER: {extracted_values.get('CUSTOMER')}")
    print(f"TYPE: {extracted_values.get('TYPE')}")
    print(f"WEIGHT: {extracted_values.get('WEIGHT')}")
    print(f"REMARK: {extracted_values.get('REMARK')}")
    print(f"ETA: {extracted_values.get('ETA')}")
    print(f"PICK UP: {extracted_values.get('PICK UP')}")

    return extracted_values

def scan_and_process_files():
    excel_paths = []
    directory = r'K:\My Drive\Company\FaCai\DO\YY'  # Updated directory path
    for filename in os.listdir(directory):
        # Skip temporary files and hidden files (those that start with ~$)
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            excel_paths.append(os.path.join(directory, filename))

    if excel_paths:
        for excel_path in excel_paths:
            print(f"Processing file: {excel_path}")
            extract_merged_cell_values(excel_path)  # Extract and print merged cell values
    else:
        print("No Excel files found in the directory!")

if __name__ == "__main__":
    scan_and_process_files()
