from openpyxl import load_workbook

def load_excel_data():
    workbook = load_workbook('do.xlsx')
    sheet = workbook['RECORD_DO']
    container_list = []
    i=0
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=6):
        row_data = []
        for cell in row:
            value = str(cell.value).replace('\n', '').strip() if cell.value else ''  # Remove newline characters
            # print(value, end=' ')
            row_data.append(value)        
        
        print()
        container_list.append(row_data)
        i+=1
    print(f'Total {i} row')    
    return container_list

container_list = load_excel_data()
# print(container_list)

i=1
for row in container_list:
    container, address, clinent, type, weight = row
    # print(container, address, clinent, type, weight)
    print(f'Row {i} {container} {address} {clinent}{type} {weight}')