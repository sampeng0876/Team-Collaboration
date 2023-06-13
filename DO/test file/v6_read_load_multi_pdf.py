import fitz
from openpyxl import load_workbook
from tkinter import Tk, Label, Button, filedialog, messagebox


class PDFCapture:
    def __init__(self, pdf_path, num_selections):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.pages = len(self.doc)
        self.page_index = 0
        self.num_selections = num_selections

    def capture_data(self):
        data = []
        x_positions = [
            (72, 126), (225, 286), (25, 268), (25, 265), (272, 428), (436, 591), (273, 427), (438, 590),
            (273, 426), (436, 589), (272, 424), (434, 508), (516, 592), (273, 508), (517, 590), (273, 426),
            (436, 590), (272, 334), (337, 365), (366, 419), (423, 492), (494, 544), (546, 591), (377, 442),
            (457, 518), (458, 517), (524, 592), (526, 592)
        ]
        y_positions = [
            (125, 135), (130, 135), (218, 263), (278, 345), (155, 172), (155, 172), (183, 196), (182, 195),
            (205, 217), (205, 218), (227, 241), (227, 241), (228, 242), (251, 265), (252, 264), (276, 289),
            (276, 289), (319, 328), (316, 328), (317, 328), (317, 329), (318, 329), (317, 330), (471, 484),
            (469, 480), (484, 498), (470, 483), (483, 501)
        ]

        for i in range(self.num_selections):
            x0, x1 = x_positions[i]
            y0, y1 = y_positions[i]
            clip_rect = fitz.Rect(x0, y0, x1, y1)
            page = self.doc.load_page(self.page_index)
            text = page.get_text("text", clip=clip_rect)
            if not text:
                text = "N/A"
            data.append(text)

        return data


def process_pdf(pdf_paths):
    # Load the workbook
    workbook = load_workbook("pdf_data.xlsx")

    # Select the sheet named 'SMART'
    sheet = workbook["SMART"]

    # Find the next available row
    next_row = sheet.max_row + 1

    for pdf_path in pdf_paths:
        # Create an instance of PDFCapture
        pdf_capture = PDFCapture(pdf_path, num_selections=28)
        captured_data = pdf_capture.capture_data()

        # Save the captured data to the next available row
        for i, data in enumerate(captured_data):
            cell = sheet.cell(row=next_row, column=i + 1)
            cell.value = data

        next_row += 1

    # Save the workbook
    workbook.save("pdf_data.xlsx")


def on_file_select():
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if file_paths:
        process_pdf(file_paths)
        messagebox.showinfo("Success", "Data captured and saved successfully!")
        root.destroy()


root = Tk()
root.title("DO Extraction")
# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the x and y coordinates for the window to be centered
x = (screen_width - 300) // 2
y = (screen_height - 200) // 2

# Set the position of the window
root.geometry(f"300x200+{x}+{y}")

label = Label(root, text="Select PDF Files")
label.pack(pady=50)

# Function to handle file selection
def select_file():
    on_file_select()

# Create a button for file selection
select_button = Button(root, text="Select Files", command=select_file)
select_button.pack()

root.mainloop()
