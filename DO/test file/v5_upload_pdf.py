import fitz
from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook("pdf_data.xlsx")

# Select the 'SMART' sheet
sheet = workbook["SMART"]

class PDFCapture:
    def __init__(self, pdf_path, num_selections):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.pages = len(self.doc)
        self.page_index = 0
        self.num_selections = num_selections

    def capture_data(self):
        data = []
        x_positions = [
            (72, 126), (225, 286), (25, 268), (25, 265), (272, 428), (436, 591), (273, 427), (438, 590),
            (273, 426), (436, 589), (272, 424), (434, 508), (516, 592), (273, 508), (517, 590), (273, 426),
            (436, 590), (272, 334), (337, 365), (366, 419), (423, 492), (494, 544), (546, 591), (377, 442),
            (457, 518), (458, 517), (524, 592), (526, 592)
        ]
        y_positions = [
            (125, 135), (130, 135), (218, 263), (278, 345), (155, 172), (155, 172), (183, 196), (182, 195),
            (205, 217), (205, 218), (227, 241), (227, 241), (228, 242), (251, 265), (252, 264), (276, 289),
            (276, 289), (319, 328), (316, 328), (317, 328), (317, 329), (318, 329), (317, 330), (471, 484),
            (469, 480), (484, 498), (470, 483), (483, 501)
        ]

        for i in range(self.num_selections):
            x0, x1 = x_positions[i]
            y0, y1 = y_positions[i]
            clip_rect = fitz.Rect(x0, y0, x1, y1)
            page = self.doc.load_page(self.page_index)
            text = page.get_text("text", clip=clip_rect)
            if not text:
                text = "N/A"
            data.append(text)

        return data


# Usage example
pdf_file = "FFAU1116554.pdf"
num_areas = 28
pdf_capture = PDFCapture(pdf_file, num_areas)
captured_data = pdf_capture.capture_data()

for i, data in enumerate(captured_data):
    print(f"Captured Data {i+1}:")
    print(data)
    print()


# # Save the captured data to the sheet starting from row 2
# for i, data in enumerate(captured_data):
#     cell = sheet.cell(row=2, column=i+1)
#     cell.value = data

# Find the next available row
next_row = sheet.max_row + 1

# Save the captured data to the next available row
for i, data in enumerate(captured_data):
    cell = sheet.cell(row=next_row, column=i+1)
    cell.value = data

# Save the workbook
workbook.save("pdf_data.xlsx")