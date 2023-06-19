from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook, load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains

chrome_options = Options()
chrome_options.add_argument('--ignore-certificate-errors')

# #chrome_options.add_argument("--disable-extensions")
# #chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--headless")
# # driver = webdriver.Chrome(options=chrome_options)


driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
# driver = webdriver.Chrome(options=chrome_options)
print('Running...')
driver.get('https://facaillc2420.yuegekeji66.cn/admin/login/login.html')
print('Loging in...')
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="admin_login"]/input[1]'))).send_keys('root')
WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="admin_login"]/input[2]'))).send_keys('123123')
driver.find_element(By.XPATH, '//*[@id="admin_login"]/button').click() #Login

WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="nav"]/li[2]/a/cite'))).click() # 点击展开 柜子汇总管理    
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="nav"]/li[2]/ul/li[1]/a/cite'))).click() #点击 柜子汇总管理
sleep(2)

def load_excel_data():
    workbook = load_workbook('do.xlsx')
    sheet = workbook['RECORD_DO']
    container_list = []
    i=0
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=6):
        row_data = []
        for cell in row:
            value = str(cell.value).replace('\n', '').strip() if cell.value else ''  # Remove newline characters
            # print(value, end=' ')
            row_data.append(value)        
        
        print()
        container_list.append(row_data)
        i+=1
    print(f'Total {i} row')    
    return container_list

container_list = load_excel_data()
print('Appending... Data')
row = 1
for data in container_list:
    container, address, clinent, type, weight = data
    print(f'Row {row} {container} {address} {clinent}{type} {weight}')
    iframe = driver.find_element(By.XPATH,'//*[@class="layui-tab-item layui-show"]/iframe') #Change iframe
    driver.switch_to.frame(iframe)

    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@class="layui-btn-container"]/button[1]'))).click() # 点击 添加数据
    sleep(2)
    iframe = driver.find_element(By.XPATH,'//*[@class="layui-layer-content"]/iframe') #Change iframe
    driver.switch_to.frame(iframe)


    driver.find_element(By.XPATH,'//*[@class="layui-tab-item layui-show"]/div[3]/div/input').send_keys(container) # 柜号
    driver.find_element(By.XPATH,'//*[@class="layui-tab-item layui-show"]/div[4]/div/input').send_keys(address) # 地址
    driver.find_element(By.XPATH,'//*[@class="layui-tab-item layui-show"]/div[5]/div/input').send_keys(clinent) # 客人
    driver.find_element(By.XPATH,'//*[@class="layui-tab-item layui-show"]/div[7]/div/input').send_keys(type) # 柜型
    driver.find_element(By.XPATH,'//*[@class="layui-tab-item layui-show"]/div[8]/div/input').send_keys(weight) # 重量
    sleep(3)
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@class="layui-form"]/div[2]/button'))).click() # 点击 增加
    sleep(3)
    row+=1


print("Finished")
driver.quit()
