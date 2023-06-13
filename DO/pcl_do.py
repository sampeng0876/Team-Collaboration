import fitz
from openpyxl import load_workbook
from tkinter import Tk, Label, Button, filedialog, messagebox


class PDFCapture:
    def __init__(self, pdf_path, num_selections):
        self.pdf_path = pdf_path
        self.doc = fitz.open(pdf_path)
        self.pages = len(self.doc)
        self.page_index = 0
        self.num_selections = num_selections

    def capture_data(self):
        data = []
        x_positions = [
            (73,118), (270,336), (22,268), (105,322), (338,361), (443,519)
        ]
        y_positions = [
            (125,135), (317,328), (288,305), (22,41), (317,329), (487,500)
        ]


        for i in range(self.num_selections):
            x0, x1 = x_positions[i]
            y0, y1 = y_positions[i]
            clip_rect = fitz.Rect(x0, y0, x1, y1)
            page = self.doc.load_page(self.page_index)
            text = page.get_text("text", clip=clip_rect)
            if not text:
                text = "N/A"
            data.append(text)

        return data


def process_pdf(pdf_paths):
    # Load the workbook
    workbook = load_workbook("/Users/champagne/Library/CloudStorage/GoogleDrive-sp.dispatchservice@gmail.com/My Drive/Company/FaCai/DO/do.xlsx")

    # Select the sheet named 'SMART'
    sheet = workbook["PCL"]

    # Find the next available row
    next_row = sheet.max_row + 1

    for pdf_path in pdf_paths:
        # Create an instance of PDFCapture
        pdf_capture = PDFCapture(pdf_path, num_selections=6)
        captured_data = pdf_capture.capture_data()

        # Save the captured data to the next available row
        for i, data in enumerate(captured_data):
            cell = sheet.cell(row=next_row, column=i + 1)
            cell.value = data

        next_row += 1

    # Save the workbook
    workbook.save("/Users/champagne/Library/CloudStorage/GoogleDrive-sp.dispatchservice@gmail.com/My Drive/Company/FaCai/DO/do.xlsx")


def on_file_select():
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
    if file_paths:
        process_pdf(file_paths)
        messagebox.showinfo("Success", "Data captured and saved successfully!")
        root.destroy()


root = Tk()
root.title("DO Extraction")
# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the x and y coordinates for the window to be centered
x = (screen_width - 300) // 2
y = (screen_height - 200) // 2

# Set the position of the window
root.geometry(f"300x200+{x}+{y}")

label = Label(root, text="Select PDF Files")
label.pack(pady=50)

# Function to handle file selection
def select_file():
    on_file_select()

# Create a button for file selection
select_button = Button(root, text="Select Files", command=select_file)
select_button.pack()

root.mainloop()
