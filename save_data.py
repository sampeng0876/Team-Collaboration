from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Vertical A2 A3 A4 A5
########################################################################
# wb = Workbook()
# ws = wb.active

# data_list = ['TRHU4481541', 'TGBU8775026', 'EGHU9499315', 'EITU1491089', 'FCGU1690407', 'CAIU4589816', 'MSCU5146165', 'MSDU7351199', 'MSDU6307268', 'DFSU3513340', 'EISU9400775', 'FDCU0033410', 'EGHU9488809', 'EITU1788267', 'TCNU2072704']

# # Starting row and column
# start_row = 2
# start_col = 1

# # Populate the data
# for index, data in enumerate(data_list):
#     cell = ws.cell(row=start_row+index, column=start_col)
#     cell.value = data

# wb.save("data.xlsx")

# print('Done')

# Horizontal A2 B2 C2 D2
########################################################################


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load the workbook
wb = load_workbook("LFD.xlsx")

# Select sheet2
ws = wb["sheet2"]

data_list = ['TRHU4481541', 'TGBU8775026', 'EGHU9499315', 'EITU1491089', 'FCGU1690407', 'CAIU4589816', 'MSCU5146165', 'MSDU7351199', 'MSDU6307268', 'DFSU3513340', 'EISU9400775', 'FDCU0033410', 'EGHU9488809', 'EITU1788267', 'TCNU2072704']

# Starting row and column
start_row = 2
start_col = 1

# Populate the data
for index, data in enumerate(data_list):
    row = start_row + (index // 3)  # Calculate the row based on the index and number of items per row
    col = start_col + (index % 3)  # Calculate the column based on the index and number of items per row

    cell = ws.cell(row=row, column=col)
    cell.value = data

# Save the modified workbook
wb.save("LFD.xlsx")


print('Done')



# Create  Sheet2
########################################################################

# wb = Workbook()

# # Remove the default sheet (Sheet1) if needed
# if "Sheet" in wb.sheetnames:
#     sheet = wb["Sheet"]
#     wb.remove(sheet)

# # Create a new sheet (Sheet2)
# ws = wb.create_sheet(title="Sheet2")

# data_list = ['TRHU4481541', 'TGBU8775026', 'EGHU9499315', 'EITU1491089', 'FCGU1690407', 'CAIU4589816', 'MSCU5146165', 'MSDU7351199', 'MSDU6307268', 'DFSU3513340', 'EISU9400775', 'FDCU0033410', 'EGHU9488809', 'EITU1788267', 'TCNU2072704']

# # Starting row and column
# start_row = 2
# start_col = 1

# # Populate the data
# for index, data in enumerate(data_list):
#     row = start_row + (index // 3)  # Calculate the row based on the index and number of items per row
#     col = start_col + (index % 3)  # Calculate the column based on the index and number of items per row

#     cell = ws.cell(row=row, column=col)
#     cell.value = data

# wb.save("data.xlsx")

# from openpyxl import Workbook

# data_list = ['', 'EMCU8377066', '143357136749', '', 'DEPARTED', 'EGL', 'EVER LASTING', '05/10/23 08:00', '05/11/23 14:16', 'COMMUNITY - OUT', 'RELEASED', 'RELEASED', '', '', '', '40/GP/96', '28065.0', '', '05/16/23 08:00', '05/16/23 10:52', '']

# workbook = Workbook()
# sheet = workbook.active

# row = 1
# column = 1
# print(data_list)
# for item in data_list:
#     sheet.cell(row=row, column=column).value = item
#     column += 1

#     if column % 4 == 0:
#         row += 1
#         column = 1

# workbook.save('data.xlsx')