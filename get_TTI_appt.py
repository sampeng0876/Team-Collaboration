from unittest import result
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from openpyxl import workbook, load_workbook
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from datetime import datetime


def read_container_list(file_path, sheet_name, column):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    row = ws.max_row
    col = ws.max_column
    print(row, '行')
    print(col, '列')
    # 获取首列从第二行开始的单元格值containers列表
    cntr = [cell.value for cell in ws['A'][1:]]
    cntr = [x for x in cntr if x is not None] #去除None值
    print(cntr)
    return cntr

def login(driver, username, password): #登录账号密码
    driver.get('https://www.ttilgb.com/main/index.do')
    # 输入用户名和密码并点击登录按钮
    driver.find_element(By.XPATH, '//*[@id="pUsrId"]').send_keys(username)
    driver.find_element(By.XPATH, '//*[@id="pUsrPwd"]').send_keys(password)
    driver.find_element(By.XPATH,'//*[@id="form"]/table/tbody/tr/td[1]/table/tbody/tr[3]/td/table[1]/tbody/tr[2]/td/table/tbody/tr[1]/td[3]/img').click() #Login
    return driver

def execute_appointment(selection): #输入 0 表示 load 列表，输入 1 表示 empty 列表
    driver = webdriver.Chrome()
    driver.maximize_window()
    actions = ActionChains(driver)
    driver = login(driver, 'P1LOGISTIC', 'p1@5418')
    cntr_list = read_container_list('刷约.xlsx', 'TTI', 1,)
    print(cntr_list)

    for ctnr_no in cntr_list:
    
        schedule_appointment(driver, ctnr_no)
    sleep(1)
    driver.quit()

def schedule_appointment(driver, ctnr_no):
    element = driver.find_element(By.XPATH,'//*[@id="nav"]/li[3]/a') #Appointment
    ActionChains(driver).move_to_element(element).perform() #鼠标悬停选择
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="nav"]/li[3]/ul/li[3]/a'))).click() #Empty Return Inquiry
    driver.find_element(By.XPATH,'//*[@id="nav"]/li[3]/ul/li[3]/a').click() #Empty Return Inquiry
    iframe = driver.find_element(By.XPATH,'//*[@id="businessView"]') #Change iframe
    driver.switch_to.frame(iframe)
    driver.find_element(By.XPATH,'//*[@id="mtyTpCd2"]').click() #Checkbox Click
    driver.find_element(By.XPATH, '//*[@id="cntrNos"]').send_keys(ctnr_no) #Empty Container
    driver.find_element(By.XPATH,'//*[@id="btnSearch"]').click() #Search
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="1"]/td[7]'))).click() #PreSelect
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="1_oprCd"]/option[13]'))).click() #Select SCAC (MSC)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="1"]/td[13]'))).click() #PreSelect
    
    # 获取当前窗口的句柄并保存
    current_window_handle = driver.current_window_handle
    
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="1"]/td[13]/img'))).click() #Select Canlendar
    # driver.find_element(By.XPATH,'//*[@id="1"]/td[13]/img').click() #Select Canlendar
    sleep(1)
    driver.find_element(By.XPATH,'//*[@id="ui-datepicker-div"]/table/tbody/tr[3]/td[4]/a').click() #Select 05/17
    
    # 查找输入框的元素并输入文本
    # 等待文本框出现并可见
    # input_box = WebDriverWait(driver, 10).until(
    #     EC.visibility_of_element_located((By.ID, "1_apntDt")))
    # input_box.send_keys("2023-05-01")
    
    # # 导入jQuery和jQuery UI库
    # jquery_script = 'https://code.jquery.com/jquery-3.6.0.min.js'
    # jquery_ui_script = 'https://code.jquery.com/ui/1.13.1/jquery-ui.min.js'
    # driver.execute_script(f'var script = document.createElement("script");script.src = "{jquery_script}";document.head.appendChild(script);')
    # driver.execute_script(f'var script = document.createElement("script");script.src = "{jquery_ui_script}";document.head.appendChild(script);')

    # # 执行JavaScript脚本
    # script = '$(".hasDatepicker").datepicker("setDate", "05-04-2023");'
    # driver.execute_script(script)
    
    # 找到日期选择器的元素并单击
    # date_input = driver.find_element(By.ID, "1_apntDt")

    # date_input.click()

    # # 选择今天的日期
    # script = '$(".hasDatepicker").datepicker("setDate", new Date());'
    # # script = '$(".hasDatepicker").datepicker("setDate", 05-03-2023);'

    # driver.execute_script(script)
    
    # # 找到日期输入框
    # date_input = driver.find_element(By.XPATH, "//td[@aria-describedby='grid1_apntDt']/input")
    # driver.execute_script("arguments[0].removeAttribute('disabled')", date_input)

    # # 清除当前日期
    # date_input.clear()

    # # 发送新日期
    # date_input.send_keys("05-03-2023")
    # element = driver.find_element(By.ID, "1_apntDt")

    # driver.execute_script("arguments[0].setAttribute('class', 'editable')", element)
    
    # # 创建一个日期对象
    # date_obj = datetime(2023, 5, 2)

    # # 将日期对象格式化为字符串格式
    # date_str = date_obj.strftime('%m-%d-%Y')

    # # 将日期字符串输入到相应的输入框中
    # input_element = driver.find_element(By.ID, '1_apntDt')
    # driver.execute_script(f"arguments[0].setAttribute('value', '{date_str}')", input_element)
    
    # # 获取当前日期并转换为字符串形式
    # date_string = datetime.now().strftime('%m-%d-%Y')

    # # 在输入框中输入日期
    # date_input = driver.find_element_by_xpath('//input[@id="1_apntDt"]')
    # date_input.send_keys(Keys.BACKSPACE * 10)  # 删除输入框中的默认值
    # date_input.send_keys(date_string)  # 输入日期
    
    # 使用JavaScript选择日期
    # date_element = driver.find_element(By.XPATH, '//*[@id="1_apntDt"]')
    # driver.execute_script("arguments[0].value = '05-02-2023';", date_element)
    
    # sleep(5)
    # WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,'//*[@id="1_tmSelCd"]'))).click() #Timeslot

    # slot_xpath = '//*[@id="1_tmSelCd"]'
    # driver.find_element(By.XPATH,slot_xpath).click() #Click Timeslot
    # slot_list = driver.find_elements(By.XPATH,slot_xpath)[0].text#.splitlines() #Read Timeslot
    # print(slot_list)

    # sleep(5)

    # # 单击确认按钮
    # # button_element = driver.find_element(By.XPATH, '//button[@id="confirm-button"]')
    # # button_element.click()

    # # 等待页面加载完成
    # wait = WebDriverWait(driver, 10)
    # result_element = wait.until(EC.presence_of_element_located((By.XPATH, '//div[@id="result"]')))

    # # 打印结果
    # print(result_element.text)
    
    # sleep(5)

    # # 找到输入框并设置为可用状态
    # date_input = driver.find_element_by_xpath('//*[@id="1_apntDt"]')
    # driver.execute_script("arguments[0].disabled = false;", date_input)

    # # 输入日期
    # date_input.send_keys('2023-05-02')
    
    # 在这里切换到 alert 弹出框并处理

    # 切换到alert弹出框并获取其文本
    sleep(3)
    alert = Alert(driver)
    alert_text = alert.text
    print(alert_text)
    if alert_text ==  'Appointments are full, check back for cancellations':
        print('没有预约')
    # 点击alert的确认按钮
    alert.accept()
    
    
    # 切换回原来的窗口
    # driver.switch_to.window(current_window_handle)

    #driver.find_element(By.XPATH,'//label[@id="btnSubmit"]').click()
    #submit = WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="btnSubmit"]')))
    driver.find_element(By.XPATH,'//*[@id="btnSubmit"]').click()
    sleep(3)
    print(f'Clicked on button submit')
    # by, buy_selector, cookies_selector = By.CSS_SELECTOR, 'button#checkout_submit', "button#onetrust-accept-btn-handler"




    # # wait for loading buy button
    # sls = wait.until(EC.presence_of_all_elements_located((by, buy_selector)))
    # if sls:
    #     # get accept cookies button element and click
    #     cookies_accept = driver.find_element_by_css_selector(cookies_selector)
    #     if isinstance(cookies_accept, WebElement):
    #         cookies_accept.click()
            
    #     # get buy button element, move to element and click
    #     buy = driver.find_element_by_css_selector(buy_selector)
    #     if isinstance(buy, WebElement) and buy.is_displayed() and buy.is_enabled():
    #         actions.move_to_element(buy).click(buy).perform()

    # driver.find_element(By.XPATH,'//*[@id="form"]/table/tbody/tr/td[1]/table/tbody/tr[1]/td/a/img').click() #Home
    sleep(1)

# chrome_options = Options()

# #chrome_options.add_argument("--disable-extensions")
# #chrome_options.add_argument("--disable-gpu")
# chrome_options.add_argument("--headless")
# # driver = webdriver.Chrome(options=chrome_options)

# cntr_list = read_container_list('刷约.xlsx', 'TTI', 1,)
# print(cntr_list)

execute_appointment(0)

print('完成')