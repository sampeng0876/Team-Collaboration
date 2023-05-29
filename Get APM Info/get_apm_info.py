from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
from time import sleep
from selenium.webdriver.chrome.options import Options
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkcalendar import DateEntry
from tkinter import *
from tkcalendar import *
import datetime as dt
import sv_ttk
import tkinter
from tkinter import ttk
import customtkinter


# Create a function to get the selected value and close the window
####################################################################################################################
def on_submit():
    global container_entry
    
    # Get container entries
    data = container_entry.get("1.0", "end-1c")
    lines = data.split("\n")
    container_list.extend(lines)
    container_entry.delete("1.0", "end")
    print("Container List:")
    print(container_list)  
    root.destroy()


# UI Window Settings
####################################################################################################################
root = tkinter.Tk()
# Set theme
# sv_ttk.use_dark_theme()
sv_ttk.use_light_theme()
# root.title("Appointment Scheduler")
# root.geometry("500x500")

root.title("Appointment Scheduler")

# Get the screen width and height
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# Calculate the x and y coordinates for the window to be centered
x = (screen_width - 210) // 2
y = (screen_height - 250) // 2

# Set the position of the window
root.geometry(f"210x250+{x}+{y}")

# List to store container entries
container_list = []

# Create a label for data entry
Label(root, text="Enter Container Number:").grid(column=1, row=1, padx=5, pady=5)
container_entry = customtkinter.CTkTextbox(root, height=160, width=200, border_width=1 ,border_color="lightgray" )
container_entry.grid(column=1, row=2, padx=5, pady=5)

# Create submit button
# style = ttk.Style()
# style.configure('Blue.TButton', foreground='blue', background='white')
submit_button = customtkinter.CTkButton(root, width=60, text="OK", command=on_submit)
submit_button.grid(column=1, row=7, padx=10, pady=5)

# Start the mainloop to display the window
root.mainloop()

# Main
####################################################################################################################
# Load the workbook
workbook = load_workbook('Get Terminal Info.xlsx')

# Select sheet2
sheet = workbook['APM']  # Change the sheet name accordingly

driver = webdriver.Chrome()
driver.maximize_window()
wait = WebDriverWait(driver, 15)
# driver.get("https://www.apmterminals.com/en/los-angeles")

driver.get("https://www.apmterminals.com/en/los-angeles/online-services/container-tracking")

sleep(2)
# Click cookies
# driver.find_element(By.XPATH, '//*[@id="main"]/div[2]/div/div[2]/button').click()
driver.find_element(By.XPATH, '//*[@id="main"]/div[3]/div/div[2]/button').click()
sleep(3)

# Click track and trace button
# containers = ['EMCU8377066',
#               'MEDU7539050',
#               'TRHU5651832',
#               'FFAU2201877',
#               'MSMU6808501',
#               'MSMU4342780',
#               'TCNU8730356',
#               'GLDU7605680']

driver.find_element(By.XPATH, '//*[@class="track-and-trace__tags"]').click()
# print(containers)
container_list = '\n'.join(container_list)
# print(container_list)

# Enter container number
driver.find_element(By.XPATH, '//div[@class="track-and-trace__tags"]/input').send_keys(container_list)

# Click submit button
driver.find_element(By.XPATH, '//*[@class="track-and-trace__submit-inner"]/button').click()


# driver.find_elements(By.XPATH, '//*[@class="trace-listing__tbody"]/tr')
# WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@class="trace-listing__tbody"]/tr')))
# driver.find_element(by='xpath', value='//*[@id="main"]/div[2]/div[1]/div[2]/div/div[8]/div/table/tbody/tr/td[5]').text
# sleep(10)

#container_id_elements = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="main"]/div[2]/div[1]/div[2]/div/div[8]/div/table/tbody/tr[1]/td')))
# table_header = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@class="trace-listing__row"]/th')))
# table_row = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, '//*[@class="trace-listing__tbody"]/tr')))

# Get data
get_data = WebDriverWait(driver, 10).until(
    EC.presence_of_all_elements_located((
    # //*[@class="trace-listing__tbody"]/tr/td # Get all data elements
    # //*[@id="main"]/div[2]/div[1]/div[2]/div/div[8]/div/table/tbody/tr[1] # 1st row data element
    # //*[@id="main"]/div[2]/div[1]/div[2]/div/div[8]/div/table/tbody/tr[2] # 2nd row data element    
    By.XPATH, '//*[@class="trace-listing__tbody"]/tr/td')))


data_list = []

# Loop through the container_id_elements and extract the text
for info in get_data:
    info_text = info.text
    print(info_text)
    data_list.append(info_text)
# print(data_list)    

# Get alt attributes Yes or No 
# alt_value = driver.find_element(By.XPATH, '//*[@ class="trace-listing__icon"]').get_attribute('alt')
get_alt = driver.find_elements(By.XPATH, '//*[@ class="trace-listing__icon"]')
alt_list = []
for info in get_alt:
    info_text = info.get_attribute('alt')
    alt_list.append(info_text)
# Print the value of "alt"
# print(alt_elements.get_attribute('alt')[0])
# print(alt_elements.get_attribute('alt')[1])
# print(alt_list)


# # V1 #################################

# Store the data to Excel file
# row = 2 # Start row
# column = 1 # Start column

# for item in data_list:

#     sheet.cell(row=row, column=column).value = item
#     column += 1

#     if column % 21 == 0:
#         row += 1
#         column = 1

# V2 ##################################

start_row = 2
start_col = 1
# Replacing values in list2
alt_list = ['Yes' if item == 'ready' else 'No' for item in alt_list]

# Iterate through the data and store it in the worksheet
for i, value in enumerate(data_list):
    col = start_col + (i % 20)
    row = start_row + (i // 20)

    # Check if the 4th, 24th, 44th, 64th, and so on elements are empty
    if i % 20 == 3 and value == '':
        value = alt_list[i // 20 % len(alt_list)]

    sheet.cell(row=row, column=col).value = value


# Save the modified workbook
workbook.save("Get Terminal Info.xlsx.xlsx")





sleep(2)
print("Done")


# Close the browser
driver.quit()